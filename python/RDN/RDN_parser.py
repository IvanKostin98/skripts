# -*- coding: utf-8 -*-
"""
@author: ivan kostin

FYI:
        Если будет необходимость сменить логотипы - ниже есть "блок лого"
            1. сохраняешь картинку в папку
            2. меняешь ссылки на логотипы
            3. меняешь высоту и ширину логотипов (операторы: height / width)
            4. take_all_division если ставишь "Да" - берет все дивизионы, ставишь "Нет" - в переменной division проставь дивизионы в '' через ,
            5. reestr_status - Да - тянет из базы данных , нет - из резервного файла Маргариты
"""
#####################################################
reestr_status = 'Да'                        #Да - из БД / Нет - из файла
poradk_nomer = '12.1'                       #это для ЦС
date_of_price = '26.03.2024'
take_all_division = 'Да'                    #Да/Нет
division = ['Юг']
#####################################################

import pandas as pd
from openpyxl import load_workbook

import openpyxl
from openpyxl.styles import (PatternFill, Alignment, Font)
import pyodbc
import time
import sys
from my_functions_only_for_import import logging_on_bd

first = logging_on_bd(username =  'Ivan')
first.ignore_warnings()

def func():
    def query(db, text):
        '''подключение к бд'''
        #Подключение к SQL (подключение только к одной датабазе)
        cnxn = pyodbc.connect(driver='{SQL Server}',
                              server='***', 
                              database=db,               
                              trusted_connection='yes')
        #Запрос можно писать в SQL Server и через слеши вставлять
        query = text
        #чтение запроса
        return pd.read_sql(query, cnxn)
    
    ############################################################################################################################
    #                                                       БЛОК ДЛЯ КОМПЕНСАЦИЙ НИЖЕ МОЦ                                      #
    ############################################################################################################################
    if take_all_division == 'Нет':
        sys.exit()
    start_comp = time.perf_counter()
    
    df_rb = query(db = 'Sandbox', text = "SELECT [Дивизион]\
                                                ,[код_1С] as [Код 1С]\
                                                ,[Клиент]\
                                                ,[L6]\
                                                ,[sku]\
                                                ,[c]\
                                                ,[по]\
                                                ,round([Цена_отгрузки_с_НДС], 3) as [Цена отгрузки с НДС]\
                                                ,round([Цена_отгрузки_без_НДС], 3) as [Цена отгрузки без НДС]\
                                                ,round([Управленческая_цена_с_НДС], 3) as [Управленческая цена с НДС]\
                                                ,round([Управленческая_цена_без_НДС], 3) as [Управленческая цена без НДС]\
                                                ,round([КУ_с_НДС], 3) as [КУ с НДС]\
                                                ,round([КУ_без_НДС], 3) as [КУ без НДС]\
                                                ,[самовывоз_процент]*100 as [самовывоз %]\
                                                ,round([самовывоз_с_НДС], 3) as [самовывоз с НДС]\
                                                ,round([самовывоз_без_НДС], 3) as [самовывоз без НДС]\
                                                ,[предоплата_процент]*100 as [предоплата %]\
                                                ,round([предоплата_с_НДС], 3) as [предоплата с НДС]\
                                                ,round([предоплата_без_НДС], 3) as [предоплата без НДС]\
                                                ,[Мероприятие]\
                                                ,[Дата_обновления] as [Дата обновления]\
                                       FROM [Sandbox].[kostin].[rdn_ku_pick_prepay_view]\
                                       WHERE year([по]) = '2024'")
    
    unique_division = df_rb['Дивизион'].unique()
    
    def centrovka(cell):
        ws[cell].font = Font(name='Arial Cyr', charset=204, family=2.0, b=True, color='C0C0C0', sz=14.0, bold=True)
        ws[cell].font = Font(bold=True)
        ws[cell].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    for i in unique_division:
        loc_df_division = df_rb.loc[df_rb['Дивизион'] == i]
        unique_distr = loc_df_division['Клиент'].unique()
        writer = pd.ExcelWriter(rf"\\***\Для всех\***\Дистрибьюторы спец.прайс\Реестр ДН нарезка по дивизионам\Компенсации дистрибьюторов персональные {i.title()}.xlsx")
        for j in unique_distr:
            print(j)
            loc_df_distr = loc_df_division.loc[loc_df_division['Клиент'] == j]
            loc_df_distr.to_excel(writer, sheet_name=j.replace('/', ''), startcol=0, startrow=12)
        writer.close()
    
    for i in unique_division:
        filename = rf"\\***\Для всех\***\Дистрибьюторы спец.прайс\Реестр ДН нарезка по дивизионам\Компенсации дистрибьюторов персональные {i.title()}.xlsx"
        df = load_workbook(filename)
        sheets_name = df.sheetnames
        for j in sheets_name:
            ws = df[j]
            ws.delete_cols(1)
            sp_use_column = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S', 'T', 'U']
            
            for i in sp_use_column:
                centrovka(f'{i}13')
                if i in ['H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S']:
                    ws.column_dimensions[i].width = 9
                for j in range(1, 13):
                    ws[i+str(j)].fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
                ws[i+str(13)].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                ws[i+str(13)].fill = PatternFill(start_color='B2D4EC', end_color='B2D4EC', fill_type='solid')
            
            ws.column_dimensions['A'].width = 20
            ws.column_dimensions['B'].width = 12
            ws.column_dimensions['C'].width = 13
            ws.column_dimensions['D'].width = 13
            ws.column_dimensions['E'].width = 44
            ws.column_dimensions['F'].width = 12
            ws.column_dimensions['G'].width = 12
            ws.column_dimensions['T'].width = 75
            ws.column_dimensions['U'].width = 12
            ws.column_dimensions.group('V','XFD', hidden=True)
                
            ws.row_dimensions[13].height = 60
            ws['G9'].value = 'КОНФИДЕНЦИАЛЬНО'
            ws['G10'].value = 'СТРОГО ВНУТРИ АСГ!'
            ws['G9'].font = Font(name='Arial Cyr', charset=204, family=2.0, b=True, color='FF0000', sz=12.0)
            ws['G10'].font = Font(name='Arial Cyr', charset=204, family=2.0, b=True, color='FF0000', sz=12.0)
            
            #блок для лого
            logo_asg = openpyxl.drawing.image.Image(r"\\***\Для всех\***\Дистрибьюторы спец.прайс\***\Python скрипты\Используем\Ценообразование\Доп файлы для работы скриптов\Confidential.jpg")
            ws.add_image(logo_asg, "F1")
    
        df.save(filename)
    end_comp = time.perf_counter()
    print(f'Компенсации отработал за: {round(end_comp - start_comp)} сек.')


first.bd_open()
try:
    func()
    first.bd_process(filename = 'RDN', processname = 'Нарезка компенсаций по дивизионам', result = 1, comments = 'загружено')
except:
    first.bd_process(filename = 'RDN', processname = 'Нарезка компенсаций по дивизионам', result = 0, comments = 'ошибка')
first.bd_close()
